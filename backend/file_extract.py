"""File extraction helpers — turn uploaded files into LLM-friendly text."""
import asyncio
import io
import csv
from pathlib import Path


async def extract_text_from_file(file_path: Path, mime_type: str = "", max_chars: int = 20000) -> str:
    """Returns extracted text or '' if not extractable."""
    def _run():
        try:
            suffix = file_path.suffix.lower()
            if suffix == ".pdf" or "pdf" in mime_type:
                from pypdf import PdfReader
                reader = PdfReader(str(file_path))
                pages = []
                for p in reader.pages[:50]:
                    try:
                        pages.append(p.extract_text() or "")
                    except Exception:
                        pages.append("")
                return "\n\n".join(pages)
            if suffix in (".csv", ".tsv"):
                text = file_path.read_text(encoding="utf-8", errors="ignore")
                # cap rows
                lines = text.splitlines()
                preview = lines[:200]
                return "\n".join(preview)
            if suffix in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(str(file_path), data_only=True, read_only=True)
                chunks = []
                for sheet_name in wb.sheetnames[:5]:
                    ws = wb[sheet_name]
                    chunks.append(f"--- Sheet: {sheet_name} ---")
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i > 150:
                            chunks.append("... (truncated)")
                            break
                        chunks.append("\t".join("" if c is None else str(c) for c in row))
                return "\n".join(chunks)
            if suffix in (".txt", ".md", ".json", ".html", ".log", ".py", ".js", ".jsx", ".ts", ".tsx", ".css", ".yaml", ".yml"):
                return file_path.read_text(encoding="utf-8", errors="ignore")
            return ""
        except Exception as e:
            return f"(extraction error: {e})"

    text = await asyncio.to_thread(_run)
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n... [truncated, {len(text)-max_chars} more chars]"
    return text
